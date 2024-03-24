from bs4 import BeautifulSoup        # import the beautifull soup packege 
import requests,openpyxl                           # import the requst and opeanpyxl packege for the store data in excel file
import pandas as pd
excel = openpyxl.Workbook()                  # store data in workbook 
from openpyxl.drawing.image import Image
import time
print(excel.sheetnames)                  # print the sheet name 

sheet = excel.active                                        
sheet.title = 'Top Rated Movies'                           #  title of the sheet print with top rated movies

print(excel.sheetnames)

             ## ## save the all columans okf the wothe name 
sheet.append(['Movie Rank','Movie Name','Year of Ralease','IMDB Rating ','your rating','imagese on imdb poster'])



try:                       #  try and except function understanding the  error 

    source = requests.get('https://www.imdb.com/chart/top/')       # get requst the on imdb movie  rating and get the access of website
    source.raise_for_status()
    print(source)


             #      get a data for understanding html form wich help of beautifull soup and html.parser 
    soup = BeautifulSoup(source.text,'html.parser') 
    
     
    
    
    
       
   
    movies = soup.find('tbody',class_='lister-list').find_all('tr')
   
    for movie in movies:
        name = movie.find('td',class_='titleColumn').a.text
        rank = movie.find('td',class_='titleColumn').get_text(strip=True).split('.')[0]        # for loop for all data in website get the my code 
        year = movie.find('td',class_='titleColumn').span.text.strip('()')
        rating = movie.find('td',class_='ratingColumn imdbRating').strong.text
        your_rating=movie.find('td',class_='ratingColumn').strong.text


        img_tag = movie.find('td', class_='posterColumn').a.img       #    get the image tag data and store the img_tag variabale  

        if img_tag:                                     #   keep the condition on image tag if the image are here this movie so get this poster with attrs heps 
            img_link =img_tag.attrs.get('src')





       
        print(rank,name,year,rating,your_rating,img_link)                 # print the variable data rank , name, year , rating 

        sheet.append([rank,name,year,rating,your_rating,img_link])     # append the sheet of all columans 

       



except Exception as e:
    print(e)

    #     save the sheet with name of imdb movie rating 

excel.save('IMDB Movie Rating.xlsx')


